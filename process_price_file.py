# -*- coding: utf-8 -*-
"""
This program converts stock price from txt file (from Dazhihui)
to xlsx file in uniform format

WARNING: Requires manually copying new sheet from new file to original file
"""

import xlrd
import xlwt
import openpyxl
import os

def process_price_file(stock_name):
    readFile = os.path.join('data', 'stock_price', ''.join([stock_name, '.txt']))
    writeFile = os.path.join('data', ''.join([stock_name, '.xlsx']))

    # workbook = xlwt.Workbook()
    # sheet = workbook.add_sheet(stock_name)
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title = stock_name)

    with open(readFile, 'r') as f:
        lines = f.readlines()
        txt_row_cnt = 3
        xlsx_row_cnt = 0
        while(True):
            try:
                line = lines[txt_row_cnt]
            except IndexError:
                break
            line = line.split('\t')
            price = line[4]
            date = line[0]
            # Year-first format, change to year-last format
            if int(date.split('/')[0]) >= 2000:
                date = date.split('/')
                y = date[0]
                m = date[1]
                d = date[2]
                date = '/'.join([d, m, y])

            # sheet.write(xlsx_row_cnt, 0, date)
            # sheet.write(xlsx_row_cnt, 1, price)
            data_pos = ''.join(['A', str(xlsx_row_cnt + 1)])
            price_pos = ''.join(['B', str(xlsx_row_cnt + 1)])
            # print("data pos: %s" % data_pos)
            ws[data_pos] = date
            ws[price_pos] = price
            xlsx_row_cnt += 1
            txt_row_cnt += 1

    # workbook.save(writeFile)
    wb.save(filename = writeFile)
    print("Stock price written in %s" % writeFile)


if __name__ == "__main__":
    process_price_file('000043')
